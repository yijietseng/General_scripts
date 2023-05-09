from pyrosetta import *
from pyrosetta.rosetta.protocols.simple_filters import ShapeComplementarityFilter
from pyrosetta.rosetta.core.simple_metrics.per_residue_metrics import PerResidueSasaMetric
from pyrosetta.rosetta.protocols.rigid import RigidBodyTransMover
from pyrosetta.rosetta.core.pose import get_chain_id_from_chain, get_jump_id_from_chain, get_jump_id_from_chain_id, get_chain_from_chain_id
from pyrosetta.rosetta.core.select.residue_selector import ResidueIndexSelector
from pyrosetta.rosetta.protocols.grafting import delete_region
import glob, argparse, os, operator, itertools
from sys import argv
import pandas as pd
from openpyxl import load_workbook

init()


def generate_initial_DataFrame(formatted_output, buried=False, chainlen=int()):
	# setting headers
	columnls = ['PDBname']
	if buried:
		if chainlen:
			for i in range(1, chainlen):
				columnls.append(f'SC{i}')
			for i in range(1, chainlen):
				columnls.append(f'BSA{i}')

		else:
			columnls += ['SC1', 'BSA1']
	else:
		if chainlen:
			for i in range(1, chainlen):
				columnls.append(f'SC{i}')
		else:
			columnls += ['SC1']
		
	df = pd.DataFrame(columns=columnls)

	# Adding data to DataFrame
	df.loc[1] = formatted_output

	return df


def write_DataFrame(logpath, df, sheetName='',folder_name_mode=''):
	# check if the log file already exists, if it does, it will create a new 
	if os.path.exists(logpath):
		with pd.ExcelWriter(logpath, engine='openpyxl', mode='a') as writer:  
			if sheetName:
				df.to_excel(writer, sheet_name=sheetName)
			else:
				'''
				If the sheet_name is not specified, it will overwrite the first worksheet.
				So to prevent this happening, below code is to automatically generate a sheet name
				If the pdb path is specified as foldernamemode, it will give the folder's name as the sheet name
				'''

				if folder_name_mode:
					newsheetName = folder_name_mode.rsplit('/',1)[0]

				else:
					# get a list with all sheets' names
					book = load_workbook(logpath)
					sheets = book.sheetnames
					# automatically generate a name for a new sheet
					for i in itertools.count(start=1):
						if f'Sheet{i}' not in sheets:
							newsheetName = f'Sheet{i}'
							break
				
				df.to_excel(writer, sheet_name=newsheetName)
		
	else:
		if sheetName:
			df.to_excel(logpath, sheet_name=sheetName)
		else:
			if folder_name_mode:
				newsheetName = folder_name_mode.rsplit('/',1)[0]
				df.to_excel(logpath, sheet_name=newsheetName)
			else:
				df.to_excel(logpath)

	print(f'Results have been saved to:{logpath}')


def get_chain_resi_ls(pose, chains=['A'], base_chain='B'):
	# Getting rosetta chainIDs
	chainIDls = []
	for i in chains:
		chainIDs = get_chain_id_from_chain(i, pose)
		chainIDls.append(chainIDs)

	# Construct residue str for each chain
	res_str_ls = []
	for chainID in chainIDls:
		res_str = ''
		chain_begin = pose.chain_begin(chainID)
		chain_end = pose.chain_end(chainID)
		for i in range(chain_begin, chain_end+1):
			resID = pose.pdb_info().pose2pdb(i).replace(' ', '')
			res_str += resID+','
		res_str = res_str[:-1]
		res_str_ls.append(res_str)

	base_chainID = get_chain_id_from_chain(base_chain, pose)
	base_chain_begin = pose.chain_begin(base_chainID)
	base_chain_end = pose.chain_end(base_chainID)
	base_res_str = ''
	for i in range(base_chain_begin, base_chain_end+1):
		base_res = pose.pdb_info().pose2pdb(i).replace(' ','')
		base_res_str += base_res+','
	base_res_str = base_res_str[:-1]


	return (res_str_ls, base_res_str)


def get_residue_ls(pose, chain='A'):

	total = pose.total_residue()
	ligstr=''
	enzstr=''
	for i in range(1,total+1):
		resi_info = pose.pdb_info().pose2pdb(i)
		resi = resi_info.replace(' ','')
		if resi[-1] == chain:
			ligstr += resi+','
		else:
			enzstr += resi+','
	ligstr = ligstr[:-1]
	enzstr = enzstr[:-1]

	return (ligstr, enzstr)


def measure_sc(pose, chain1str='', chain2str=''):
	sc = ShapeComplementarityFilter()

	# if a ligand is identified, it measure the ligand's sc. Otherwise, it measures global sc
	if chain1str and chain2str:
		sc.residues1(chain1str)
		sc.residues2(chain2str)
	
	sc_score = sc.score(pose)
	
	return sc_score


def ligand_sc(PDBname, chain='A', logpath='', buried=False):
	pose = pose_from_pdb(PDBname)
		
	ligstr, enzstr = get_residue_ls(pose, chain)
	sc_score = measure_sc(pose, ligstr, enzstr)
	if buried:
		buried_sa = measure_buried_SA(pose)
		if logpath:
			f = open(logpath,'a+')
			f.write('PDBFN\tSC\tBuried_sa\n')
			f.write(f'{PDBname}\t{sc_score}\t{buried_sa}\n')
			f. close()
			print(f'log is saved to {logpath}')

		print(f'SC score for {PDBname} is: {sc_score}')
		print(f'Buried surface area of the ligand is: {buried_sa}')
	else:
		if logpath:
			f = open(logpath,'a+')
			f.write('PDBFN\tSC\n')
			f.write(f'{PDBname}\t{sc_score}\n')
			f. close()
			print(f'log is saved to {logpath}')

		print(f'SC score for {PDBname} is: {sc_score}')


def multi_sc(PDBpath, logpath='', buried=False, focus='', sheetName=''):
	'''
	Measures SC for multiple structures.
	'''
	if isinstance(PDBpath,list):
		PDB = PDBpath

	elif os.path.isdir(PDBpath):
		if PDBpath[-1] != '/' and '*' not in PDBpath:
			PDBpath += '/'
			# identified PDB list in a specified directory
			PDB = glob.glob(PDBpath+'*.pdb')
		elif PDBpath[-1] != '/' and '*' in PDBpath:
			PDB = glob.glob(PDBpath)

	elif '.txt' in PDBpath:
		# or extract PDBs from a txt file
		with open(PDBpath) as f:
			lines = f.read().splitlines()
		for line in lines:
			index = lines.index(line)
			# if a line state something like ex: ./*.pdb, then it will grab all pdb file names and will compile them into a big list
			if '*' in line:
				lines[index] = glob.glob(line)
			else:
				lines[index] = line.split()
		PDB = list(itertools.chain(*lines))

	#print(PDB)
	# organize SCs into a list
	outputls = []
	for i in PDB:
		# calculate SC and BSA for each individual PDB
		sc_bsa_output_ls = single_sc(i, buried=buried, focus=focus)
		outputls.append(sc_bsa_output_ls)
	#outputls.sort(reverse=True)
	#print(outputls)
	
	df = pd.DataFrame()
	pd.set_option('display.max_rows', 500)
	pd.set_option('display.max_columns', 500)
	pd.set_option('display.width', 1000)
	# generate and append results in to a single dataframe
	df = pd.DataFrame()
	pd.set_option('display.max_rows', 500)
	pd.set_option('display.max_columns', 500)
	pd.set_option('display.width', 1000)
	for output in outputls:
		df = df.append(output, ignore_index=True)
	df = df.sort_index( axis=1)

	print('\n\n\nRun completed!!!!')
	print(df)	

	if logpath:
		write_DataFrame(logpath, df,  sheetName=sheetName)
		print(f'Results have saved to:{logpath}')

def single_sc(PDBname, logpath='', focus='', sheetName='', buried=False):
	
	# loading PDB
	pose = pose_from_pdb(PDBname)

	# get total chain number 
	total_chain = pose.num_chains()

	
	# focus is a str of the basechain, ex: 'B'
	if focus:
		# getting the chains besides the base chain
		basechain = focus
		chain1ls = []
		num_chains = pose.num_chains()
		for i in range(1, num_chains+1):
			chain1ls.append(get_chain_from_chain_id(i, pose))
		chain1ls.remove(basechain)

		# getting the residue string for each chain	
		chain1strls, basechainstr = get_chain_resi_ls(pose, chains=chain1ls, base_chain=basechain)
		#print(f'chain1ls = {chain1ls}')

		# calculating and formatting the sc ouput 
		sc_score_output= {}
		num = 1
		for i in chain1strls:
			sc_score = measure_sc(pose, chain1str=i, chain2str=basechainstr)
			sc_score_output['SC'+str(num)] = sc_score
			num += 1
		
		# calculating the buried surface area and output into a text file 
		if buried:
			buried_sa = measure_buried_SA(pose, basechain=basechain, excluding_chain=chain1ls)
			# formatting PDBname into a dict and append the SC result to it
			formatted_output = {'APDBname':PDBname}
			formatted_output.update(sc_score_output)
			formatted_output.update(buried_sa)
			if logpath:
				# generate a dataframe for result and save 
				df = pd.DataFrame()
				df = df.append(formatted_output, ignore_index=True)
				df = df.sort_index( axis=1)
				write_DataFrame(logpath, df, sheetName)

				print(f'SC score for {PDBname} is: {"  ".join(str(s) for s in list(sc_score_output.values()))}')
				print(f'Calculated buried surface area is: {"  ".join(str(s) for s in list(buried_sa.values()))}')

				return formatted_output

			else:
				print(f'SC score for {PDBname} is: {"  ".join(str(s) for s in list(sc_score_output.values()))}')
				print(f'Calculated buried surface area is: {"  ".join(str(s) for s in list(buried_sa.values()))}')

				return formatted_output
		
		else:
			formatted_output = PDBname.split()+sc_score_output
			if logpath:
				df = pd.DataFrame()
				df = df.append(formatted_output, ignore_index=True)
				df = df.sort_index( axis=1)
				write_DataFrame(logpath, df, sheetName)
				print(f'SC score for {PDBname} is: {"  ".join(str(s) for s in list(sc_score_output.values()))}')
				
				return formatted_output
			
			else:
				print(f'SC score for {PDBname} is: {"  ".join(str(s) for s in list(sc_score_output.values()))}')

				return formatted_output

	else:
		sc_score = measure_sc(pose)
		if buried:
			buried_sa = measure_buried_SA(pose)
			formatted_output = PDBname.split()+sc_score_output+buried_sa
			if logpath:
				df = pd.DataFrame()
				df = df.append(formatted_output, ignore_index=True)
				df = df.sort_index( axis=1)
				write_DataFrame(logpath, df, sheetName)
				print(f'SC score for {PDBname} is: {"  ".join(str(s) for s in list(sc_score_output.values()))}')
				print(f'Calculated buried surface area is: {"  ".join(str(s) for s in list(buried_sa.values()))}')
			
				return formatted_output
			
			else:
				print(f'SC score for {PDBname} is: {"  ".join(str(s) for s in list(sc_score_output.values()))}')
				print(f'Calculated buried surface area is: {"  ".join(str(s) for s in list(buried_sa.values()))}')
			
				return formatted_output

		else:
			formatted_output = PDBname.split()+sc_score_output
			if logpath:
				df = pd.DataFrame()
				df = df.append(formatted_output, ignore_index=True)
				df = df.sort_index( axis=1)
				write_DataFrame(logpath, df, sheetName)
				print(f'SC score for {PDBname} is: {"  ".join(str(s) for s in list(sc_score_output.values()))}')
			
				return formatted_output
			
			else:
				print(f'SC score for {PDBname} is: {"  ".join(str(s) for s in list(sc_score_output.values()))}')
			
				return formatted_output


def measure_buried_SA(pose, basechain='A', excluding_chain=[]):
	'''
	This function will measure the buried surface area by measuring the dSASA. 
	If excluding chain is specified, it create tempory poses to delete the chains specified in the excluding chain list,
	and measure only the individual dSASA against the basechain. If the excluding_chain is not specified, it will assume
	there are two subunits and will measure the dSASA by only translating the two subunits apart. 
	'''
	
	sasa_metric = PerResidueSasaMetric()

	# measure SASA for individual chain identified in the excluding_chain
	if excluding_chain:
		#print(excluding_chain)
		SASAdic = {}
		for i, chains in enumerate(excluding_chain):
			current_excluding_chain = excluding_chain[:i] + excluding_chain[i+1:]
			#print(current_excluding_chain)
			#print(chains)
			# select the chains to be measured
			chain1, chain2 = get_chain_resi_ls(pose, chains, basechain)
			#print(f'{chain1}\n\n')
			#print(f'{chain2}\n\n')
			chainstr = chain1[0]+','+chain2
			#print(f'combine: {chainstr}')
			chainSelector = ResidueIndexSelector(chainstr)
			sasa_metric.set_residue_selector(chainSelector)
			# creating initial pose where all chains but the selected ones are translated away
			pose_init = Pose()
			pose_init.assign(pose)
			#$pose_init = RBTranslate(pose_init, basechain=basechain, excluding_chain=current_excluding_chain, init=True)
			pose_init = delete_chain(pose_init, excluding_chain=current_excluding_chain)
			# Creating a translated pose
			pose_t = Pose()
			pose_t.assign(pose_init)
			pose_t = RBTranslate(pose_t,basechain=basechain,step_size=1500)
			#pose_init.dump_pdb(f'test{i}.pdb')
			#pose_t.dump_pdb(f'trans{i}.pdb')

			#calculating SASA
			resi_sasa1 = sasa_metric.calculate(pose_init)
			resi_sasa2 = sasa_metric.calculate(pose_t)


			# Getting the SASA for each residue
			resi_sasa1 = sorted(resi_sasa1.items(), key=operator.itemgetter(1), reverse=False)
			resi_sasa2 = sorted(resi_sasa2.items(), key=operator.itemgetter(1), reverse=False)
			relist1, salist1 = zip(*resi_sasa1)
			relist1, salist2 = zip(*resi_sasa2)

			# Calculating buried surface area
			SA1 = sum(salist1)
			SA2 = sum(salist2)

			SASAdic['BSA'+str(i+1)] = SA2 - SA1

		return SASAdic
		

	else:
		# Creating a translated pose
		pose_t = Pose()
		pose_t.assign(pose)
		pose_t = RBTranslate(pose_t, basechain=basechain)

		# Getting the SASA for each residue
		resi_sasa1 = sasa_metric.calculate(pose)
		resi_sasa2 = sasa_metric.calculate(pose_t)

		# Formatting the SASA
		resi_sasa1 = sorted(resi_sasa1.items(), key=operator.itemgetter(1), reverse=False)
		resi_sasa2 = sorted(resi_sasa2.items(), key=operator.itemgetter(1), reverse=False)
		relist1, salist1 = zip(*resi_sasa1)
		relist1, salist2 = zip(*resi_sasa2)
		#print(resi_sasa1)
		#print(resi_sasa2)

		# Calculating buried surface area
		SA1 = sum(salist1)
		SA2 = sum(salist2)
		return SA2 - SA1


def delete_chain(pose, excluding_chain=['C']):

	if len(excluding_chain) == 1:
		print('\nOnly one excluding chain detected, returning same pose as the initial pose\n')
		return pose 
	else:
		# Reversing excluding chain, because deleting chain bottom up can keep the protein length consistant and avoid deleting the wrong chain
		excluding_chain.reverse()
		
		# Getting the residue pose numbers for begining and ending of the chain as tuples
		resi_ls = []
		for i in excluding_chain:
			chainID = get_chain_id_from_chain(i, pose)
			begining = pose.chain_begin(chainID)
			ending = pose.chain_end(chainID)
			resi_ls.append((begining, ending))

		# Deleting the chain
		for begin, end in resi_ls:
			delete_region(pose, begin, end)
		
		return pose


def RBTranslate(pose, basechain='A', step_size=500):

	basechainID = get_chain_id_from_chain(basechain, pose)
	try:
		base_jumpID = get_jump_id_from_chain(basechain, pose)
	except RuntimeError:
		base_jumpID = get_jump_id_from_chain_id(basechainID+1, pose)
		new_basechain = get_chain_from_chain_id(basechainID+1, pose)
		print(f'Resetting jump to Chain {new_basechain}\n')
	# perform rigid body translation
	trans_mover = RigidBodyTransMover(pose, base_jumpID) 
	trans_mover.step_size(step_size) # 500 Ang is the default
	trans_mover.apply(pose)
	return pose


#-------------------------Below section is definitions of input options------------------------------------------

parser = argparse.ArgumentParser(description='Measure the SC score for 1 or multiple structures, it can also measure the SC for just the Ligand')
parser.add_argument('-b',
					'--buried', 
					action='store_true',
					help='Calculate buried surface area')
parser.add_argument('-f',
					'--focus',
					nargs='+',
					help='Focus mode is to calculate SC at specific interface(s). First argument indicates')
parser.add_argument("-i", 
					'--input',
					nargs='+',
					required=True,
					help='input PDB file path/name')
parser.add_argument('-m',
					'--multi', 
					action='store_true',
					help='Measure SC for multiple structures')
parser.add_argument('-l',
					'--ligand',
					nargs='?',
					const='A',
					type=str,
					help='Measure SC of ligand, argument is the ChainID of the ligand, defualt = A')
parser.add_argument('-r',
					'--record',
					nargs='+',
					type=str,
					help='Record the result into an xlsx file or not. It takes two arguments: 1. file name/path to create into or to read from (w/ w/o .xlsx, it will force the output into an xlsx file) 2. WorkSheet name')
parser.add_argument('-s',
					'--single', 
					action='store_true',
					help='Measure SC for a single structure')


args = parser.parse_args()

#---------------------------------------------Section ends----------------------------------------------------


def main():
	if len(args.input) > 1:
		PDB = args.input
	else:
		PDB = args.input[0]
		print(PDB)
		print(type(PDB))
		
	if args.record:
		logpath = args.record[0]
		#formating logpath
		if '.xlsx' not in logpath:
			logpath += '.xlsx'
		if len(args.record) == 2:
			sheetName = args.record[1]
		elif len(args.record) == 1:
			sheetName = ''
	else:
		logpath = ''
		sheetName = ''

	if args.single:
		if os.path.isdir(PDB):
			raise ValueError('Please input a path to a single PDB file')
		else:
			if args.buried:
				if args.focus:
					focus = args.focus[0]
					single_sc(PDB, logpath, buried=True, focus=focus, sheetName=sheetName)
				else:
					single_sc(PDB, logpath, buried=True, sheetName=sheetName)
			else:
				if args.focus:
					focus = args.focus[0]
					single_sc(PDB, logpath, focus=focus, sheetName=sheetName)
				else:
					single_sc(PDB, logpath, sheetName=sheetName)
	elif args.multi:
		if isinstance(PDB, list) or os.path.isdir(PDB) or '.txt' in PDB:
			if args.buried:
				if args.focus:
					focus = args.focus[0]
					multi_sc(PDB, logpath, buried=True, focus=focus, sheetName=sheetName)
				else:
					multi_sc(PDB, logpath, buried=True, sheetName=sheetName)
			else:
				if args.focus:
					focus = args.focus[0]
					multi_sc(PDB, logpath, focus=focus, sheetName=sheetName)
				else:
					multi_sc(PDB, logpath, sheetName=sheetName)
		else:
			raise ValueError('Please input a path to multiple PDB files or input a txt file with all PDB paths')
	
	elif args.ligand:
		if os.path.isdir(PDB):
			raise ValueError('Please input a path to a single PDB file with ligand')
		else:
			if args.buried:
				ligand_sc(PDB, args.ligand[0],logpath, buried=True)
			else:
				ligand_sc(PDB, args.ligand[0],logpath)


if __name__ == '__main__':
	main()
